import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Border, Side, NamedStyle, PatternFill, Font, Alignment
from datetime import datetime


def schreibe_excel(
    saison_file="2025.json",
    wochentag="freitag",
    spieler_file="spieler.json",
    output_file="spielplan1.xlsx",
):
    saison_path = Path(saison_file)
    spieler_path = Path(spieler_file)

    if not saison_path.exists() or not spieler_path.exists():
        raise FileNotFoundError("⚠️ Saison- oder Spielerdatei fehlt!")

    saison = json.loads(saison_path.read_text(encoding="utf-8"))
    spieler_stamm = {sp["id"]: sp for sp in json.loads(spieler_path.read_text(encoding="utf-8"))}

    if "groups" not in saison or wochentag not in saison["groups"]:
        raise KeyError(f"⚠️ Gruppe {wochentag} nicht gefunden in {saison_file}")

    gruppe = saison["groups"][wochentag]
    verteilung = gruppe.get("verteilung", {})
    gruppe_spieler = list(gruppe["players"].keys())

    jahr = saison.get("jahr", "????")

    # Workbook anlegen
    wb = Workbook()
    ws = wb.active
    ws.title = f"Gruppe {gruppe['wochentag']}"

    # Styles
    date_style = NamedStyle(name="date_style")
    date_style.number_format = "DD.MM.YYYY"

    thin_border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )
    thick_side = Side(style="thick", color="000000")

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # Schriftarten
    header_font = Font(size=14, bold=True)   # Datum + Spieler-Header
    title_font = Font(size=18, bold=True)    # Überschrift oben

    # Titelzeilen
    breite = len(gruppe_spieler) + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=breite)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=breite)

    ws["A1"] = f"{gruppe['wochentag']} – Runde {jahr}"
    ws["A2"] = f"Platz {gruppe['platz']} – von {gruppe['startzeit']} bis {gruppe['endzeit']}"

    for cell in (ws["A1"], ws["A2"]):
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Spieler-Header (Zeile 5)
    ws.cell(row=5, column=1, value="Datum").font = header_font
    for col, pid in enumerate(gruppe_spieler, start=2):
        name = spieler_stamm.get(pid, {}).get("name", pid)
        cell = ws.cell(row=5, column=col, value=name)
        cell.font = header_font

    # Daten eintragen (Zeilen ab 6)
    for row, (date, assigned_players) in enumerate(verteilung.items(), start=6):
        try:
            date_obj = datetime.strptime(date, "%d.%m.%Y")
        except ValueError:
            date_obj = date
        date_cell = ws.cell(row=row, column=1, value=date_obj)
        if isinstance(date_obj, datetime):
            date_cell.style = date_style

        for col, pid in enumerate(gruppe_spieler, start=2):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border  # erst dünner Rahmen

            if pid in assigned_players:
                cell.value = 1
            elif date in gruppe["players"][pid].get("nicht_moeglich", []):
                cell.fill = red_fill

    # Außenrahmen fett (NACH den dünnen Rahmen)
    max_row = ws.max_row
    max_col = ws.max_column
    for row in range(5, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            b = cell.border

            left = b.left or thin_border.left
            right = b.right or thin_border.right
            top = b.top or thin_border.top
            bottom = b.bottom or thin_border.bottom

            if col == 1:
                left = thick_side
            if col == max_col:
                right = thick_side
            if row == 5:
                top = thick_side
            if row == max_row:
                bottom = thick_side

            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Spaltenbreiten anpassen
    # Spalte A (Datum): auto_width + 5
    max_len_date = max(len(str(ws.cell(row=r, column=1).value)) for r in range(5, max_row + 1))
    ws.column_dimensions["A"].width = max_len_date + 5

    # Alle Spieler-Spalten gleich breit
    max_len = 0
    for col in range(2, max_col + 1):
        col_letter = ws.cell(row=5, column=col).column_letter
        text = str(ws.cell(row=5, column=col).value)
        if len(text) > max_len:
            max_len = len(text)
    for col in range(2, max_col + 1):
        col_letter = ws.cell(row=5, column=col).column_letter
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_file)
    print(f"✅ Excel-Datei gespeichert: {output_file}")


if __name__ == "__main__":
    schreibe_excel("2025.json", "freitag", "spieler.json", "spielplan.xlsx")
