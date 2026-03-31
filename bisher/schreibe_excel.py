import json
from openpyxl import Workbook
from openpyxl.styles import Border, Side, NamedStyle
from datetime import datetime


def schreibe_excel(json_file='24-Freitag.json', output_file='spielplan.xlsx'):
    # Load JSON data
    with open(json_file, 'r') as file:
        data = json.load(file)

    players = data['players']
    verteilung = data['verteilung']

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Create a date style
    date_style = NamedStyle(name='date_style')
    date_style.number_format = 'DD.MM.YYYY'

    # Write headers (player names)
    for col, player in enumerate(players, start=2):
        ws.cell(row=5, column=col, value=player['name'])

    # Define red border style
    red_border = Border(left=Side(style='thin', color='FF0000'),
                        right=Side(style='thin', color='FF0000'),
                        top=Side(style='thin', color='FF0000'),
                        bottom=Side(style='thin', color='FF0000'))

    # Fill in the matrix
    for row, (date, assigned_players) in enumerate(verteilung.items(), start=6):
        # Convert date string to datetime object
        date_obj = datetime.strptime(date, '%d.%m.%Y')
        # Write the date in the first column and apply the date style
        date_cell = ws.cell(row=row, column=1, value=date_obj)
        date_cell.style = date_style

        for col, player in enumerate(players, start=2):
            cell = ws.cell(row=row, column=col)

            if player['name'] in assigned_players:
                cell.value = 1
            elif date in player['nicht_verfuegbare_termine']:
                # cell.value = 0
                cell.border = red_border
            # else:
            #     cell.value = 0

    # Save the workbook
    wb.save(output_file)


if __name__ == "__main__":
    schreibe_excel()